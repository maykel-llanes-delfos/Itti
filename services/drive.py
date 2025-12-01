"""Servicio de Google Drive"""

import io
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime

from googleapiclient.discovery import build
from googleapiclient.http import (
    MediaFileUpload,
    MediaIoBaseDownload,
    MediaInMemoryUpload,
)
from googleapiclient.errors import HttpError
import openpyxl
import pandas as pd

from config.settings import AppConfig
from models.schemas import ArchivoCliente, DriveFileChange, ExcelData


class GoogleDriveService:
    """Servicio para interactuar con Google Drive"""

    def __init__(self, credentials, config: AppConfig):
        self.service = build("drive", "v3", credentials=credentials)
        self.config = config
        self._carpetas_cache = {}  # Cache de carpetas {nombre: folder_id}

    def crear_carpeta(
        self, nombre_carpeta: str, parent_id: Optional[str] = None
    ) -> str:
        """
        Crea una carpeta en Drive (funciÃ³n genÃ©rica)

        Args:
            nombre_carpeta: Nombre de la carpeta
            parent_id: ID de la carpeta padre (opcional)

        Returns:
            ID de la carpeta creada
        """
        folder_metadata = {
            "name": nombre_carpeta,
            "mimeType": "application/vnd.google-apps.folder",
        }

        if parent_id:
            folder_metadata["parents"] = [parent_id]
        elif self.config.drive_root_folder_id:
            folder_metadata["parents"] = [self.config.drive_root_folder_id]

        try:
            folder = (
                self.service.files()
                .create(body=folder_metadata, fields="id, name, webViewLink")
                .execute()
            )

            print(f"âœ… Carpeta creada: {nombre_carpeta}")
            print(f"   ID: {folder['id']}")
            print(f"   Link: {folder.get('webViewLink', 'N/A')}")

            return folder["id"]

        except HttpError as error:
            print(f"âŒ Error creando carpeta: {error}")
            raise

    def subir_archivo(self, archivo: ArchivoCliente, folder_id: str) -> str:
        """
        Sube un archivo a Drive (funciÃ³n genÃ©rica)

        Args:
            archivo: Datos del archivo a subir
            folder_id: ID de la carpeta destino

        Returns:
            ID del archivo subido
        """
        archivo.validar_contenido()

        file_metadata = {"name": archivo.nombre_destino, "parents": [folder_id]}

        try:
            # Subir desde archivo local
            if archivo.ruta_local:
                media = MediaFileUpload(
                    archivo.ruta_local, mimetype=archivo.mime_type, resumable=True
                )
            # Subir desde bytes en memoria
            else:
                media = MediaInMemoryUpload(
                    archivo.contenido_bytes, mimetype=archivo.mime_type, resumable=True
                )

            file = (
                self.service.files()
                .create(
                    body=file_metadata, media_body=media, fields="id, name, webViewLink"
                )
                .execute()
            )

            print(f"  ğŸ“ Archivo subido: {file['name']}")
            print(f"     ID: {file['id']}")

            return file["id"]

        except HttpError as error:
            print(f"âŒ Error subiendo archivo: {error}")
            raise

    def subir_multiples_archivos(
        self, archivos: List[ArchivoCliente], folder_id: str
    ) -> Dict[str, str]:
        """
        Sube mÃºltiples archivos a una carpeta

        Args:
            archivos: Lista de archivos a subir
            folder_id: ID de la carpeta destino

        Returns:
            Diccionario {nombre_archivo: file_id}
        """
        archivos_subidos = {}

        for archivo in archivos:
            file_id = self.subir_archivo(archivo, folder_id)
            archivos_subidos[archivo.nombre_destino] = file_id

        return archivos_subidos

    def crear_carpeta_con_archivos(
        self,
        nombre_carpeta: str,
        archivos: List[ArchivoCliente],
        parent_id: Optional[str] = None,
    ) -> tuple[str, Dict[str, str]]:
        """
        Crea carpeta y sube archivos (funciÃ³n genÃ©rica completa)

        Args:
            nombre_carpeta: Nombre de la carpeta a crear
            archivos: Lista de archivos a subir
            parent_id: ID de carpeta padre (opcional)

        Returns:
            Tupla (folder_id, dict_archivos_subidos)
        """
        print(
            f"\nğŸ“¦ Creando carpeta '{nombre_carpeta}' "
            f"con {len(archivos)} archivos..."
        )

        # Crear carpeta
        folder_id = self.crear_carpeta(nombre_carpeta, parent_id)

        # Subir archivos
        archivos_subidos = self.subir_multiples_archivos(archivos, folder_id)

        print(f"âœ… Proceso completado: " f"{len(archivos_subidos)} archivos subidos")

        return folder_id, archivos_subidos

    def buscar_carpeta_por_nombre(
        self, nombre: str, parent_id: Optional[str] = None
    ) -> Optional[str]:
        """Busca una carpeta por nombre"""
        query = (
            f"name='{nombre}' and "
            f"mimeType='application/vnd.google-apps.folder' and "
            f"trashed=false"
        )

        if parent_id:
            query += f" and '{parent_id}' in parents"
        elif self.config.drive_root_folder_id:
            query += f" and '{self.config.drive_root_folder_id}' in parents"

        try:
            results = (
                self.service.files()
                .list(q=query, spaces="drive", fields="files(id, name, webViewLink)")
                .execute()
            )

            files = results.get("files", [])

            if files:
                print(
                    f"ğŸ“‚ Carpeta encontrada: {files[0]['name']} "
                    f"(ID: {files[0]['id']})"
                )
                return files[0]["id"]
            else:
                print(f"ğŸ“‚ Carpeta '{nombre}' no encontrada")
                return None

        except HttpError as error:
            print(f"âŒ Error buscando carpeta: {error}")
            return None

    def listar_archivos_excel(
        self, folder_id: str, modified_after: Optional[datetime] = None
    ) -> List[DriveFileChange]:
        """
        Lista archivos Excel en una carpeta

        Args:
            folder_id: ID de la carpeta
            modified_after: Solo archivos modificados despuÃ©s de esta fecha

        Returns:
            Lista de archivos Excel encontrados
        """
        query = (
            f"'{folder_id}' in parents and "
            "(mimeType='application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet' or "
            "mimeType='application/vnd.ms-excel' or "
            "mimeType='application/vnd.google-apps.spreadsheet') and "
            "trashed=false"
        )

        if modified_after:
            timestamp = modified_after.isoformat() + "Z"
            query += f" and modifiedTime > '{timestamp}'"

        try:
            results = (
                self.service.files()
                .list(
                    q=query,
                    spaces="drive",
                    fields="files(id, name, modifiedTime, mimeType, "
                    "parents, webViewLink)",
                    orderBy="modifiedTime desc",
                )
                .execute()
            )

            files = results.get("files", [])

            cambios = []
            for file in files:
                cambio = DriveFileChange(
                    file_id=file["id"],
                    file_name=file["name"],
                    modified_time=datetime.fromisoformat(
                        file["modifiedTime"].replace("Z", "+00:00")
                    ),
                    mime_type=file["mimeType"],
                    parent_folder=(file["parents"][0] if file.get("parents") else ""),
                    web_view_link=file.get("webViewLink"),
                )
                cambios.append(cambio)

            return cambios

        except HttpError as error:
            print(f"âŒ Error listando archivos: {error}")
            return []

    def descargar_archivo(self, file_id: str, destino: str) -> None:
        """Descarga un archivo de Drive a disco"""
        try:
            # Para Google Sheets, exportar como Excel
            file_metadata = (
                self.service.files().get(fileId=file_id, fields="mimeType").execute()
            )

            if file_metadata["mimeType"] == "application/vnd.google-apps.spreadsheet":
                request = self.service.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet",
                )
            else:
                request = self.service.files().get_media(fileId=file_id)

            Path(destino).parent.mkdir(parents=True, exist_ok=True)

            with io.FileIO(destino, "wb") as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    if status:
                        print(f"  â¬‡ï¸  Descarga " f"{int(status.progress() * 100)}%")

            print(f"ğŸ’¾ Archivo descargado: {destino}")

        except HttpError as error:
            print(f"âŒ Error descargando archivo: {error}")
            raise

    def leer_excel_desde_drive(self, file_id: str) -> ExcelData:
        """
        Lee un archivo Excel directamente desde Drive

        Args:
            file_id: ID del archivo Excel

        Returns:
            ExcelData con toda la informaciÃ³n del archivo
        """
        try:
            # Obtener metadata
            file_metadata = (
                self.service.files()
                .get(fileId=file_id, fields="name, modifiedTime, mimeType")
                .execute()
            )

            # Descargar contenido
            if file_metadata["mimeType"] == "application/vnd.google-apps.spreadsheet":
                request = self.service.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet",
                )
            else:
                request = self.service.files().get_media(fileId=file_id)

            # Leer en memoria
            file_bytes = io.BytesIO()
            downloader = MediaIoBaseDownload(file_bytes, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            file_bytes.seek(0)

            # Parsear Excel con openpyxl
            wb = openpyxl.load_workbook(file_bytes, data_only=True)

            data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Convertir a lista de diccionarios
                rows = []
                headers = [cell.value for cell in ws[1]]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for header, value in zip(headers, row):
                        row_dict[header] = value
                    rows.append(row_dict)

                data[sheet_name] = rows

            excel_data = ExcelData(
                file_id=file_id,
                file_name=file_metadata["name"],
                sheet_names=wb.sheetnames,
                data=data,
                modified_time=datetime.fromisoformat(
                    file_metadata["modifiedTime"].replace("Z", "+00:00")
                ),
            )

            print(f"ğŸ“Š Excel leÃ­do: {excel_data.file_name}")
            print(f"   Hojas: {', '.join(excel_data.sheet_names)}")
            print(f"   Total filas: " f"{sum(len(rows) for rows in data.values())}")

            return excel_data

        except HttpError as error:
            print(f"âŒ Error leyendo Excel: {error}")
            raise

    def actualizar_excel_en_drive(
        self, file_id: str, dataframes: Dict[str, pd.DataFrame]
    ) -> None:
        """
        Actualiza un archivo Excel en Drive

        Args:
            file_id: ID del archivo a actualizar
            dataframes: Diccionario {sheet_name: DataFrame}
        """
        try:
            # Crear Excel en memoria
            excel_bytes = io.BytesIO()

            with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            excel_bytes.seek(0)

            # Subir a Drive
            media = MediaInMemoryUpload(
                excel_bytes.read(),
                mimetype="application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet",
                resumable=True,
            )

            self.service.files().update(fileId=file_id, media_body=media).execute()

            print(f"âœ… Excel actualizado en Drive (ID: {file_id})")

        except HttpError as error:
            print(f"âŒ Error actualizando Excel: {error}")
            raise

    def obtener_o_crear_carpeta(
        self, nombre_carpeta: str, parent_id: Optional[str] = None
    ) -> str:
        """
        Obtiene el ID de una carpeta existente o la crea si no existe

        Args:
            nombre_carpeta: Nombre de la carpeta
            parent_id: ID de la carpeta padre (opcional)

        Returns:
            ID de la carpeta (existente o nueva)
        """
        # Buscar carpeta existente
        folder_id = self.buscar_carpeta_por_nombre(nombre_carpeta, parent_id)

        if folder_id:
            print(f"âœ… Usando carpeta existente: {nombre_carpeta}")
            return folder_id

        # Si no existe, crearla
        print(f"ğŸ“ Creando nueva carpeta: {nombre_carpeta}")
        return self.crear_carpeta(nombre_carpeta, parent_id)

    def listar_todas_carpetas(
        self, parent_id: Optional[str] = None, actualizar_cache: bool = True
    ) -> Dict[str, str]:
        """
        Lista todas las carpetas en un directorio y actualiza el cache

        Args:
            parent_id: ID de la carpeta padre (opcional)
            actualizar_cache: Si True, actualiza el cache con los resultados

        Returns:
            Diccionario {nombre_carpeta: folder_id}
        """
        query = f"mimeType='application/vnd.google-apps.folder' and " f"trashed=false"

        if parent_id:
            query += f" and '{parent_id}' in parents"
        elif self.config.drive_root_folder_id:
            query += f" and '{self.config.drive_root_folder_id}' in parents"

        try:
            carpetas = {}
            page_token = None

            while True:
                results = (
                    self.service.files()
                    .list(
                        q=query,
                        spaces="drive",
                        fields="nextPageToken, files(id, name)",
                        pageToken=page_token,
                        pageSize=100,
                    )
                    .execute()
                )

                files = results.get("files", [])

                for file in files:
                    carpetas[file["name"]] = file["id"]

                    # Actualizar cache
                    if actualizar_cache:
                        cache_key = (
                            f"{parent_id or self.config.drive_root_folder_id}:"
                            f"{file['name']}"
                        )
                        self._carpetas_cache[cache_key] = file["id"]

                page_token = results.get("nextPageToken")
                if not page_token:
                    break

            print(f"ğŸ“‚ Encontradas {len(carpetas)} carpetas")
            return carpetas

        except HttpError as error:
            print(f"âŒ Error listando carpetas: {error}")
            return {}

    def procesar_clientes_desde_excel(
        self,
        excel_data: "ExcelData",
        columna_nombre: str = "Nombre",
        crear_carpetas: bool = True,
        parent_id: Optional[str] = None,
    ) -> Dict[str, str]:
        """
        Procesa clientes desde Excel y crea/obtiene sus carpetas

        Args:
            excel_data: Datos del Excel
            columna_nombre: Nombre de la columna con nombres de clientes
            crear_carpetas: Si True, crea carpetas para clientes nuevos
            parent_id: ID de carpeta padre (opcional)

        Returns:
            Diccionario {nombre_cliente: folder_id}
        """
        print(f"\nğŸ“Š Procesando clientes desde Excel: {excel_data.file_name}")

        # Obtener todos los nombres de clientes (sin duplicados)
        clientes = set()

        for sheet_name, rows in excel_data.data.items():
            for row in rows:
                if columna_nombre in row and row[columna_nombre]:
                    nombre = str(row[columna_nombre]).strip()
                    if nombre:
                        clientes.add(nombre)

        print(f"ğŸ“‹ Encontrados {len(clientes)} clientes Ãºnicos")

        # Listar carpetas existentes para optimizar bÃºsquedas
        print("\nğŸ” Listando carpetas existentes...")
        self.listar_todas_carpetas(parent_id, actualizar_cache=True)

        # Procesar cada cliente
        carpetas_clientes = {}

        for i, nombre_cliente in enumerate(sorted(clientes), 1):
            print(f"\n[{i}/{len(clientes)}] Procesando: {nombre_cliente}")

            if crear_carpetas:
                folder_id = self.obtener_o_crear_carpeta(nombre_cliente, parent_id)
                carpetas_clientes[nombre_cliente] = folder_id
            else:
                folder_id = self.buscar_carpeta_por_nombre(nombre_cliente, parent_id)
                if folder_id:
                    carpetas_clientes[nombre_cliente] = folder_id

        print(f"\nâœ… Procesados {len(carpetas_clientes)} clientes")
        return carpetas_clientes
