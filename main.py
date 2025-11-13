"""
Sistema de Gestión de Google Drive y Gmail con Service Accounts
Requisitos: pip install google-api-python-client google-auth pydantic openpyxl pandas python-dotenv gspread
"""

import os
import io
import base64
import time
import json
from pathlib import Path
from typing import Optional, List, Dict, Any, Literal
from datetime import datetime
from enum import Enum

from pydantic import BaseModel, Field, field_validator
from pydantic_settings import BaseSettings
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import (
    MediaFileUpload,
    MediaIoBaseDownload,
    MediaInMemoryUpload,
)
from googleapiclient.errors import HttpError
import openpyxl
import pandas as pd


# ============================================================================
# CONFIGURACIÓN Y MODELOS PYDANTIC
# ============================================================================


class AuthMode(str, Enum):
    """Modo de autenticación"""

    PERSONAL = "personal"  # Service Account para uso personal
    WORKSPACE = "workspace"  # Service Account con domain-wide delegation


class GoogleScopes:
    """Scopes de Google API"""

    DRIVE_FULL = "https://www.googleapis.com/auth/drive"
    GMAIL_READONLY = "https://www.googleapis.com/auth/gmail.readonly"
    GMAIL_MODIFY = "https://www.googleapis.com/auth/gmail.modify"


class AppConfig(BaseSettings):
    """Configuración de la aplicación desde variables de entorno"""

    # Modo de autenticación
    auth_mode: AuthMode = Field(default=AuthMode.PERSONAL)

    # Service Account
    service_account_file: str = Field(
        default="service-account.json", description="Archivo JSON de la service account"
    )

    # Para Workspace con domain-wide delegation
    delegated_user_email: Optional[str] = Field(
        default=None, description="Email del usuario a suplantar (solo para Workspace)"
    )

    # Configuración Gmail
    gmail_filter_subject: Optional[str] = Field(default=None)
    gmail_filter_from: Optional[str] = Field(default=None)
    gmail_filter_label: Optional[str] = Field(default=None)
    gmail_check_interval: int = Field(default=60)

    # Configuración Drive
    drive_root_folder_id: Optional[str] = Field(default=None)
    drive_check_interval: int = Field(default=300)

    # Rutas locales
    local_download_path: str = Field(default="./downloads")
    local_temp_path: str = Field(default="./temp")

    class Config:
        env_file = ".env"
        env_file_encoding = "utf-8"


class Cliente(BaseModel):
    """Modelo de datos para un cliente"""

    nombre: str = Field(..., min_length=1)
    apellido1: str = Field(..., min_length=1)
    apellido2: str = Field(..., min_length=1)

    @field_validator("nombre", "apellido1", "apellido2")
    @classmethod
    def validar_mayusculas(cls, v: str) -> str:
        return v.strip().upper()

    @property
    def nombre_carpeta(self) -> str:
        return f"{self.nombre} {self.apellido1} {self.apellido2}"


class ArchivoCliente(BaseModel):
    """Modelo para archivos a subir"""

    ruta_local: Optional[str] = Field(None, description="Ruta local del archivo")
    contenido_bytes: Optional[bytes] = Field(None, description="Contenido en bytes")
    nombre_destino: str = Field(..., description="Nombre del archivo en Drive")
    mime_type: str = Field(default="application/octet-stream")

    @field_validator("ruta_local")
    @classmethod
    def validar_archivo_existe(cls, v: Optional[str]) -> Optional[str]:
        if v and not Path(v).exists():
            raise ValueError(f"El archivo no existe: {v}")
        return v

    def validar_contenido(self):
        """Valida que exista ruta_local o contenido_bytes"""
        if not self.ruta_local and not self.contenido_bytes:
            raise ValueError("Debe proporcionar ruta_local o contenido_bytes")


class EmailAttachment(BaseModel):
    """Modelo para archivos adjuntos de email"""

    filename: str
    mime_type: str
    data: bytes
    size: int = Field(gt=0)


class EmailMessage(BaseModel):
    """Modelo para mensajes de email"""

    id: str
    thread_id: str
    subject: str
    sender: str
    date: datetime
    has_attachments: bool = False
    attachments: List[EmailAttachment] = Field(default_factory=list)


class DriveFileChange(BaseModel):
    """Modelo para cambios en archivos de Drive"""

    file_id: str
    file_name: str
    modified_time: datetime
    mime_type: str
    parent_folder: str
    web_view_link: Optional[str] = None


class ExcelData(BaseModel):
    """Modelo para datos leídos de Excel"""

    file_id: str
    file_name: str
    sheet_names: List[str]
    data: Dict[str, List[Dict[str, Any]]]  # {sheet_name: [rows]}
    modified_time: datetime


# ============================================================================
# SERVICIO DE AUTENTICACIÓN GOOGLE
# ============================================================================


class GoogleAuthService:
    """Servicio de autenticación con Service Accounts"""

    def __init__(self, config: AppConfig):
        self.config = config
        self.scopes = [
            GoogleScopes.DRIVE_FULL,
            GoogleScopes.GMAIL_READONLY,
            GoogleScopes.GMAIL_MODIFY,
        ]

    def get_credentials(self, for_service: Literal["drive", "gmail"] = "drive"):
        """
        Obtiene credenciales de Service Account

        Args:
            for_service: "drive" o "gmail"

        Returns:
            Credenciales configuradas
        """
        if not os.path.exists(self.config.service_account_file):
            raise FileNotFoundError(
                f"Archivo de service account no encontrado: {self.config.service_account_file}"
            )

        # Cargar credenciales base
        credentials = service_account.Credentials.from_service_account_file(
            self.config.service_account_file, scopes=self.scopes
        )

        # Para Workspace con domain-wide delegation
        if self.config.auth_mode == AuthMode.WORKSPACE:
            if not self.config.delegated_user_email:
                raise ValueError(
                    "Para modo WORKSPACE debes configurar DELEGATED_USER_EMAIL"
                )

            # Delegar credenciales al usuario
            credentials = credentials.with_subject(self.config.delegated_user_email)
            print(f"🔐 Usando delegación para: {self.config.delegated_user_email}")
        else:
            print(f"🔐 Usando Service Account en modo PERSONAL")

        return credentials


# ============================================================================
# SERVICIO DE GOOGLE DRIVE
# ============================================================================


class GoogleDriveService:
    """Servicio para interactuar con Google Drive"""

    def __init__(self, credentials, config: AppConfig):
        self.service = build("drive", "v3", credentials=credentials)
        self.config = config

    def crear_carpeta(
        self, nombre_carpeta: str, parent_id: Optional[str] = None
    ) -> str:
        """
        Crea una carpeta en Drive (función genérica)

        Args:
            nombre_carpeta: Nombre de la carpeta
            parent_id: ID de la carpeta padre (opcional)

        Returns:
            ID de la carpeta creada
        """
        folder_metadata = {
            "name": nombre_carpeta,
            "mimeType": "application/vnd.google-apps.folder",
        }

        if parent_id:
            folder_metadata["parents"] = [parent_id]
        elif self.config.drive_root_folder_id:
            folder_metadata["parents"] = [self.config.drive_root_folder_id]

        try:
            folder = (
                self.service.files()
                .create(body=folder_metadata, fields="id, name, webViewLink")
                .execute()
            )

            print(f"✅ Carpeta creada: {nombre_carpeta}")
            print(f"   ID: {folder['id']}")
            print(f"   Link: {folder.get('webViewLink', 'N/A')}")

            return folder["id"]

        except HttpError as error:
            print(f"❌ Error creando carpeta: {error}")
            raise

    def subir_archivo(self, archivo: ArchivoCliente, folder_id: str) -> str:
        """
        Sube un archivo a Drive (función genérica)

        Args:
            archivo: Datos del archivo a subir
            folder_id: ID de la carpeta destino

        Returns:
            ID del archivo subido
        """
        archivo.validar_contenido()

        file_metadata = {"name": archivo.nombre_destino, "parents": [folder_id]}

        try:
            # Subir desde archivo local
            if archivo.ruta_local:
                media = MediaFileUpload(
                    archivo.ruta_local, mimetype=archivo.mime_type, resumable=True
                )
            # Subir desde bytes en memoria
            else:
                media = MediaInMemoryUpload(
                    archivo.contenido_bytes, mimetype=archivo.mime_type, resumable=True
                )

            file = (
                self.service.files()
                .create(
                    body=file_metadata, media_body=media, fields="id, name, webViewLink"
                )
                .execute()
            )

            print(f"  📁 Archivo subido: {file['name']}")
            print(f"     ID: {file['id']}")

            return file["id"]

        except HttpError as error:
            print(f"❌ Error subiendo archivo: {error}")
            raise

    def subir_multiples_archivos(
        self, archivos: List[ArchivoCliente], folder_id: str
    ) -> Dict[str, str]:
        """
        Sube múltiples archivos a una carpeta

        Args:
            archivos: Lista de archivos a subir
            folder_id: ID de la carpeta destino

        Returns:
            Diccionario {nombre_archivo: file_id}
        """
        archivos_subidos = {}

        for archivo in archivos:
            file_id = self.subir_archivo(archivo, folder_id)
            archivos_subidos[archivo.nombre_destino] = file_id

        return archivos_subidos

    def crear_carpeta_con_archivos(
        self,
        nombre_carpeta: str,
        archivos: List[ArchivoCliente],
        parent_id: Optional[str] = None,
    ) -> tuple[str, Dict[str, str]]:
        """
        Crea carpeta y sube archivos (función genérica completa)

        Args:
            nombre_carpeta: Nombre de la carpeta a crear
            archivos: Lista de archivos a subir
            parent_id: ID de carpeta padre (opcional)

        Returns:
            Tupla (folder_id, dict_archivos_subidos)
        """
        print(
            f"\n📦 Creando carpeta '{nombre_carpeta}' con {len(archivos)} archivos..."
        )

        # Crear carpeta
        folder_id = self.crear_carpeta(nombre_carpeta, parent_id)

        # Subir archivos
        archivos_subidos = self.subir_multiples_archivos(archivos, folder_id)

        print(f"✅ Proceso completado: {len(archivos_subidos)} archivos subidos")

        return folder_id, archivos_subidos

    def buscar_carpeta_por_nombre(
        self, nombre: str, parent_id: Optional[str] = None
    ) -> Optional[str]:
        """Busca una carpeta por nombre"""
        query = f"name='{nombre}' and mimeType='application/vnd.google-apps.folder' and trashed=false"

        if parent_id:
            query += f" and '{parent_id}' in parents"
        elif self.config.drive_root_folder_id:
            query += f" and '{self.config.drive_root_folder_id}' in parents"

        try:
            results = (
                self.service.files()
                .list(q=query, spaces="drive", fields="files(id, name, webViewLink)")
                .execute()
            )

            files = results.get("files", [])

            if files:
                print(
                    f"📂 Carpeta encontrada: {files[0]['name']} (ID: {files[0]['id']})"
                )
                return files[0]["id"]
            else:
                print(f"📂 Carpeta '{nombre}' no encontrada")
                return None

        except HttpError as error:
            print(f"❌ Error buscando carpeta: {error}")
            return None

    def listar_archivos_excel(
        self, folder_id: str, modified_after: Optional[datetime] = None
    ) -> List[DriveFileChange]:
        """
        Lista archivos Excel en una carpeta

        Args:
            folder_id: ID de la carpeta
            modified_after: Solo archivos modificados después de esta fecha

        Returns:
            Lista de archivos Excel encontrados
        """
        query = (
            f"'{folder_id}' in parents and "
            "(mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or "
            "mimeType='application/vnd.ms-excel' or "
            "mimeType='application/vnd.google-apps.spreadsheet') and "
            "trashed=false"
        )

        if modified_after:
            timestamp = modified_after.isoformat() + "Z"
            query += f" and modifiedTime > '{timestamp}'"

        try:
            results = (
                self.service.files()
                .list(
                    q=query,
                    spaces="drive",
                    fields="files(id, name, modifiedTime, mimeType, parents, webViewLink)",
                    orderBy="modifiedTime desc",
                )
                .execute()
            )

            files = results.get("files", [])

            cambios = []
            for file in files:
                cambio = DriveFileChange(
                    file_id=file["id"],
                    file_name=file["name"],
                    modified_time=datetime.fromisoformat(
                        file["modifiedTime"].replace("Z", "+00:00")
                    ),
                    mime_type=file["mimeType"],
                    parent_folder=file["parents"][0] if file.get("parents") else "",
                    web_view_link=file.get("webViewLink"),
                )
                cambios.append(cambio)

            return cambios

        except HttpError as error:
            print(f"❌ Error listando archivos: {error}")
            return []

    def descargar_archivo(self, file_id: str, destino: str) -> None:
        """Descarga un archivo de Drive a disco"""
        try:
            # Para Google Sheets, exportar como Excel
            file_metadata = (
                self.service.files().get(fileId=file_id, fields="mimeType").execute()
            )

            if file_metadata["mimeType"] == "application/vnd.google-apps.spreadsheet":
                request = self.service.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                request = self.service.files().get_media(fileId=file_id)

            Path(destino).parent.mkdir(parents=True, exist_ok=True)

            with io.FileIO(destino, "wb") as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    if status:
                        print(f"  ⬇️  Descarga {int(status.progress() * 100)}%")

            print(f"💾 Archivo descargado: {destino}")

        except HttpError as error:
            print(f"❌ Error descargando archivo: {error}")
            raise

    def leer_excel_desde_drive(self, file_id: str) -> ExcelData:
        """
        Lee un archivo Excel directamente desde Drive

        Args:
            file_id: ID del archivo Excel

        Returns:
            ExcelData con toda la información del archivo
        """
        try:
            # Obtener metadata
            file_metadata = (
                self.service.files()
                .get(fileId=file_id, fields="name, modifiedTime, mimeType")
                .execute()
            )

            # Descargar contenido
            if file_metadata["mimeType"] == "application/vnd.google-apps.spreadsheet":
                request = self.service.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                request = self.service.files().get_media(fileId=file_id)

            # Leer en memoria
            file_bytes = io.BytesIO()
            downloader = MediaIoBaseDownload(file_bytes, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            file_bytes.seek(0)

            # Parsear Excel con openpyxl
            wb = openpyxl.load_workbook(file_bytes, data_only=True)

            data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Convertir a lista de diccionarios
                rows = []
                headers = [cell.value for cell in ws[1]]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for header, value in zip(headers, row):
                        row_dict[header] = value
                    rows.append(row_dict)

                data[sheet_name] = rows

            excel_data = ExcelData(
                file_id=file_id,
                file_name=file_metadata["name"],
                sheet_names=wb.sheetnames,
                data=data,
                modified_time=datetime.fromisoformat(
                    file_metadata["modifiedTime"].replace("Z", "+00:00")
                ),
            )

            print(f"📊 Excel leído: {excel_data.file_name}")
            print(f"   Hojas: {', '.join(excel_data.sheet_names)}")
            print(f"   Total filas: {sum(len(rows) for rows in data.values())}")

            return excel_data

        except HttpError as error:
            print(f"❌ Error leyendo Excel: {error}")
            raise

    def actualizar_excel_en_drive(
        self, file_id: str, dataframes: Dict[str, pd.DataFrame]
    ) -> None:
        """
        Actualiza un archivo Excel en Drive

        Args:
            file_id: ID del archivo a actualizar
            dataframes: Diccionario {sheet_name: DataFrame}
        """
        try:
            # Crear Excel en memoria
            excel_bytes = io.BytesIO()

            with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            excel_bytes.seek(0)

            # Subir a Drive
            media = MediaInMemoryUpload(
                excel_bytes.read(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=True,
            )

            self.service.files().update(fileId=file_id, media_body=media).execute()

            print(f"✅ Excel actualizado en Drive (ID: {file_id})")

        except HttpError as error:
            print(f"❌ Error actualizando Excel: {error}")
            raise


# ============================================================================
# SERVICIO DE GMAIL
# ============================================================================


class GmailService:
    """Servicio para interactuar con Gmail"""

    def __init__(self, credentials, config: AppConfig):
        self.service = build("gmail", "v1", credentials=credentials)
        self.config = config

    def _construir_query(self) -> str:
        """Construye query de búsqueda"""
        queries = ["has:attachment"]

        if self.config.gmail_filter_subject:
            queries.append(f"subject:{self.config.gmail_filter_subject}")

        if self.config.gmail_filter_from:
            queries.append(f"from:{self.config.gmail_filter_from}")

        if self.config.gmail_filter_label:
            queries.append(f"label:{self.config.gmail_filter_label}")

        return " ".join(queries)

    def buscar_correos(
        self, max_results: int = 10, unread_only: bool = True
    ) -> List[EmailMessage]:
        """Busca correos según filtros configurados"""
        query = self._construir_query()

        if unread_only:
            query += " is:unread"

        print(f"🔍 Buscando correos: {query}")

        try:
            results = (
                self.service.users()
                .messages()
                .list(userId="me", q=query, maxResults=max_results)
                .execute()
            )

            messages = results.get("messages", [])

            email_messages = []
            for msg in messages:
                email_msg = self._procesar_mensaje(msg["id"])
                if email_msg:
                    email_messages.append(email_msg)

            return email_messages

        except HttpError as error:
            print(f"❌ Error buscando correos: {error}")
            return []

    def _procesar_mensaje(self, msg_id: str) -> Optional[EmailMessage]:
        """Procesa un mensaje y extrae información"""
        try:
            msg = (
                self.service.users()
                .messages()
                .get(userId="me", id=msg_id, format="full")
                .execute()
            )

            headers = msg["payload"]["headers"]

            subject = next(
                (h["value"] for h in headers if h["name"].lower() == "subject"), ""
            )
            sender = next(
                (h["value"] for h in headers if h["name"].lower() == "from"), ""
            )
            date_str = next(
                (h["value"] for h in headers if h["name"].lower() == "date"), ""
            )

            from email.utils import parsedate_to_datetime

            date = parsedate_to_datetime(date_str) if date_str else datetime.now()

            attachments = self._extraer_adjuntos(msg)

            return EmailMessage(
                id=msg["id"],
                thread_id=msg["threadId"],
                subject=subject,
                sender=sender,
                date=date,
                has_attachments=len(attachments) > 0,
                attachments=attachments,
            )

        except HttpError as error:
            print(f"❌ Error procesando mensaje: {error}")
            return None

    def _extraer_adjuntos(self, msg: Dict[str, Any]) -> List[EmailAttachment]:
        """Extrae archivos adjuntos de un mensaje"""
        attachments = []

        def procesar_parte(part: Dict[str, Any]):
            if "filename" in part and part["filename"]:
                if "body" in part and "attachmentId" in part["body"]:
                    try:
                        attachment_id = part["body"]["attachmentId"]
                        attachment = (
                            self.service.users()
                            .messages()
                            .attachments()
                            .get(userId="me", messageId=msg["id"], id=attachment_id)
                            .execute()
                        )

                        data = base64.urlsafe_b64decode(
                            attachment["data"].encode("UTF-8")
                        )

                        att = EmailAttachment(
                            filename=part["filename"],
                            mime_type=part["mimeType"],
                            data=data,
                            size=attachment["size"],
                        )
                        attachments.append(att)
                    except HttpError as error:
                        print(f"⚠️  Error extrayendo adjunto: {error}")

            if "parts" in part:
                for subpart in part["parts"]:
                    procesar_parte(subpart)

        if "parts" in msg["payload"]:
            for part in msg["payload"]["parts"]:
                procesar_parte(part)

        return attachments

    def marcar_como_leido(self, msg_id: str) -> None:
        """Marca un mensaje como leído"""
        try:
            self.service.users().messages().modify(
                userId="me", id=msg_id, body={"removeLabelIds": ["UNREAD"]}
            ).execute()
        except HttpError as error:
            print(f"⚠️  Error marcando como leído: {error}")

    def extraer_excel_adjuntos(self, correo: EmailMessage) -> List[EmailAttachment]:
        """
        Extrae solo los adjuntos Excel de un correo

        Args:
            correo: Mensaje de email

        Returns:
            Lista de adjuntos Excel
        """
        return [
            att
            for att in correo.attachments
            if att.filename.lower().endswith((".xlsx", ".xls"))
        ]


# ============================================================================
# JOBS Y PROCESOS
# ============================================================================


class EmailProcessorJob:
    """Job para procesar correos entrantes con adjuntos Excel"""

    def __init__(
        self,
        gmail_service: GmailService,
        drive_service: GoogleDriveService,
        config: AppConfig,
    ):
        self.gmail_service = gmail_service
        self.drive_service = drive_service
        self.config = config

    def procesar_correos_nuevos(
        self,
        guardar_local: bool = True,
        subir_a_drive: bool = True,
        folder_id_drive: Optional[str] = None,
    ) -> List[tuple[EmailMessage, List[str]]]:
        """
        Procesa correos nuevos y maneja adjuntos Excel

        Args:
            guardar_local: Si True, guarda adjuntos localmente
            subir_a_drive: Si True, sube adjuntos a Drive
            folder_id_drive: ID de carpeta Drive destino

        Returns:
            Lista de tuplas (correo, [file_ids_en_drive])
        """
        print("\n📧 Procesando correos nuevos...")

        correos = self.gmail_service.buscar_correos(unread_only=True)

        if not correos:
            print("  No hay correos nuevos")
            return []

        print(f"  Encontrados {len(correos)} correos nuevos")

        resultados = []

        for correo in correos:
            print(f"\n  📨 Procesando: {correo.subject}")
            print(f"     De: {correo.sender}")
            print(f"     Fecha: {correo.date}")

            excel_attachments = self.gmail_service.extraer_excel_adjuntos(correo)

            if not excel_attachments:
                print("     ⚠️  No se encontraron adjuntos Excel")
                continue

            file_ids = []

            for att in excel_attachments:
                # Guardar localmente
                if guardar_local:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{timestamp}_{att.filename}"
                    destino = os.path.join(self.config.local_download_path, filename)

                    Path(destino).parent.mkdir(parents=True, exist_ok=True)
                    with open(destino, "wb") as f:
                        f.write(att.data)

                    print(f"     💾 Guardado local: {destino}")

                # Subir a Drive
                if subir_a_drive:
                    if not folder_id_drive:
                        folder_id_drive = self.config.drive_root_folder_id

                    if folder_id_drive:
                        archivo = ArchivoCliente(
                            contenido_bytes=att.data,
                            nombre_destino=att.filename,
                            mime_type=att.mime_type,
                        )

                        file_id = self.drive_service.subir_archivo(
                            archivo, folder_id_drive
                        )
                        file_ids.append(file_id)
                    else:
                        print("     ⚠️  No hay folder_id_drive configurado")

            # Marcar como leído
            self.gmail_service.marcar_como_leido(correo.id)

            resultados.append((correo, file_ids))

        print(f"\n✅ Procesados {len(resultados)} correos")
        return resultados

    def ejecutar_loop(
        self,
        guardar_local: bool = True,
        subir_a_drive: bool = True,
        folder_id_drive: Optional[str] = None,
    ):
        """Ejecuta el job en loop continuo"""
        print("🚀 Iniciando EmailProcessorJob...")
        print(f"   Intervalo: {self.config.gmail_check_interval}s")

        while True:
            try:
                self.procesar_correos_nuevos(
                    guardar_local=guardar_local,
                    subir_a_drive=subir_a_drive,
                    folder_id_drive=folder_id_drive,
                )
            except Exception as e:
                print(f"❌ Error procesando correos: {e}")

            time.sleep(self.config.gmail_check_interval)


class DriveMonitorJob:
    """Job para monitorear cambios en archivos Excel en Drive"""

    def __init__(self, drive_service: GoogleDriveService, config: AppConfig):
        self.drive_service = drive_service
        self.config = config
        self.last_check: Optional[datetime] = None

    def procesar_cambios(
        self, folder_id: str, callback_on_change: Optional[callable] = None
    ) -> List[ExcelData]:
        """
        Detecta y procesa cambios en archivos Excel

        Args:
            folder_id: ID de la carpeta a monitorear
            callback_on_change: Función a ejecutar cuando hay cambios
                                Debe aceptar ExcelData como parámetro

        Returns:
            Lista de archivos Excel modificados
        """
        print(f"\n📊 Monitoreando cambios en Excel (Carpeta: {folder_id})...")

        cambios = self.drive_service.listar_archivos_excel(folder_id, self.last_check)

        if not cambios:
            print("  No hay cambios")
            self.last_check = datetime.now()
            return []

        print(f"  ✨ Detectados {len(cambios)} cambios")

        excel_data_list = []

        for cambio in cambios:
            print(f"\n  📝 Archivo modificado: {cambio.file_name}")
            print(f"     ID: {cambio.file_id}")
            print(f"     Modificado: {cambio.modified_time}")
            print(f"     Link: {cambio.web_view_link}")

            try:
                # Leer Excel desde Drive
                excel_data = self.drive_service.leer_excel_desde_drive(cambio.file_id)
                excel_data_list.append(excel_data)

                # Ejecutar callback si existe
                if callback_on_change:
                    print(f"     ⚡ Ejecutando callback personalizado...")
                    callback_on_change(excel_data)

            except Exception as e:
                print(f"     ❌ Error procesando Excel: {e}")

        self.last_check = datetime.now()
        return excel_data_list

    def ejecutar_loop(
        self, folder_id: str, callback_on_change: Optional[callable] = None
    ):
        """Ejecuta el job en loop continuo"""
        print("🚀 Iniciando DriveMonitorJob...")
        print(f"   Intervalo: {self.config.drive_check_interval}s")

        while True:
            try:
                self.procesar_cambios(folder_id, callback_on_change)
            except Exception as e:
                print(f"❌ Error monitoreando cambios: {e}")

            time.sleep(self.config.drive_check_interval)


# ============================================================================
# UTILIDADES Y HELPERS
# ============================================================================


class PlaceholderGenerator:
    """Generador de archivos placeholder para testing"""

    @staticmethod
    def crear_imagen_placeholder(
        nombre: str, ancho: int = 800, alto: int = 600, color: str = "blue"
    ) -> ArchivoCliente:
        """
        Crea una imagen placeholder en memoria

        Args:
            nombre: Nombre del archivo
            ancho: Ancho de la imagen
            alto: Alto de la imagen
            color: Color de la imagen

        Returns:
            ArchivoCliente con la imagen en memoria
        """
        try:
            from PIL import Image, ImageDraw, ImageFont

            # Crear imagen
            img = Image.new("RGB", (ancho, alto), color=color)
            draw = ImageDraw.Draw(img)

            # Agregar texto
            text = f"PLACEHOLDER\n{nombre}"
            bbox = draw.textbbox((0, 0), text)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]

            position = ((ancho - text_width) // 2, (alto - text_height) // 2)
            draw.text(position, text, fill="white")

            # Convertir a bytes
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            img_bytes.seek(0)

            return ArchivoCliente(
                contenido_bytes=img_bytes.read(),
                nombre_destino=nombre,
                mime_type="image/png",
            )

        except ImportError:
            # Si PIL no está disponible, crear archivo de texto
            print("⚠️  PIL no disponible, creando placeholder de texto")
            contenido = f"PLACEHOLDER IMAGE: {nombre}".encode("utf-8")

            return ArchivoCliente(
                contenido_bytes=contenido, nombre_destino=nombre, mime_type="text/plain"
            )

    @staticmethod
    def crear_excel_placeholder(nombre: str = "placeholder.xlsx") -> ArchivoCliente:
        """
        Crea un archivo Excel placeholder

        Args:
            nombre: Nombre del archivo

        Returns:
            ArchivoCliente con el Excel en memoria
        """
        # Crear DataFrame de ejemplo
        df = pd.DataFrame(
            {
                "ID": range(1, 11),
                "Nombre": [f"Item {i}" for i in range(1, 11)],
                "Cantidad": [i * 10 for i in range(1, 11)],
                "Precio": [i * 100.5 for i in range(1, 11)],
            }
        )

        # Guardar en memoria
        excel_bytes = io.BytesIO()
        df.to_excel(excel_bytes, index=False, sheet_name="Datos")
        excel_bytes.seek(0)

        return ArchivoCliente(
            contenido_bytes=excel_bytes.read(),
            nombre_destino=nombre,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ============================================================================
# EJEMPLOS DE USO Y CALLBACKS PERSONALIZADOS
# ============================================================================


def ejemplo_callback_validacion(excel_data: ExcelData):
    """
    Ejemplo de callback para validación de datos
    Se ejecuta cada vez que se detecta un cambio en Excel
    """
    print(f"\n🔍 VALIDANDO: {excel_data.file_name}")

    for sheet_name, rows in excel_data.data.items():
        print(f"\n  📋 Hoja: {sheet_name}")
        print(f"     Total filas: {len(rows)}")

        if not rows:
            print("     ⚠️  Hoja vacía")
            continue

        # Ejemplo: Validar que no haya valores nulos en columnas críticas
        columnas_requeridas = ["ID", "Nombre"]  # Ajustar según tu caso

        for col in columnas_requeridas:
            if col in rows[0]:
                valores_nulos = sum(1 for row in rows if not row.get(col))
                if valores_nulos > 0:
                    print(f"     ⚠️  Columna '{col}': {valores_nulos} valores nulos")
                else:
                    print(f"     ✅ Columna '{col}': OK")

        # Ejemplo: Validar rangos numéricos
        if "Precio" in rows[0]:
            precios = [row.get("Precio", 0) for row in rows if row.get("Precio")]
            if precios:
                min_precio = min(precios)
                max_precio = max(precios)
                print(
                    f"     💰 Rango de precios: ${min_precio:.2f} - ${max_precio:.2f}"
                )

                if min_precio < 0:
                    print(f"     ❌ ERROR: Hay precios negativos")


def ejemplo_callback_actualizar_bd(excel_data: ExcelData):
    """
    Ejemplo de callback para actualizar base de datos
    """
    print(f"\n💾 ACTUALIZANDO BASE DE DATOS desde: {excel_data.file_name}")

    # Aquí iría tu código de actualización de BD
    # Ejemplo con SQLAlchemy:
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    
    engine = create_engine('postgresql://user:pass@localhost/dbname')
    Session = sessionmaker(bind=engine)
    session = Session()
    
    for sheet_name, rows in excel_data.data.items():
        for row in rows:
            # Actualizar o insertar en BD
            registro = MiModelo(**row)
            session.merge(registro)
    
    session.commit()
    """

    print(
        f"     ✅ {sum(len(rows) for rows in excel_data.data.values())} registros procesados"
    )


def ejemplo_callback_generar_reporte(excel_data: ExcelData):
    """
    Ejemplo de callback para generar reportes automáticos
    """
    print(f"\n📊 GENERANDO REPORTE desde: {excel_data.file_name}")

    for sheet_name, rows in excel_data.data.items():
        if not rows:
            continue

        # Convertir a DataFrame para análisis
        df = pd.DataFrame(rows)

        print(f"\n  📈 Análisis de '{sheet_name}':")
        print(f"     Total registros: {len(df)}")

        # Estadísticas de columnas numéricas
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            print(f"\n     Columnas numéricas:")
            for col in numeric_cols:
                print(f"       - {col}:")
                print(f"         Media: {df[col].mean():.2f}")
                print(f"         Suma: {df[col].sum():.2f}")


# ============================================================================
# FUNCIONES PRINCIPALES DE USO
# ============================================================================


def setup_servicios(config: AppConfig) -> tuple[GoogleDriveService, GmailService]:
    """
    Configura e inicializa los servicios de Google

    Args:
        config: Configuración de la aplicación

    Returns:
        Tupla (drive_service, gmail_service)
    """
    print("\n" + "=" * 60)
    print("INICIALIZANDO SERVICIOS DE GOOGLE")
    print("=" * 60)

    # Autenticación
    auth_service = GoogleAuthService(config)

    # Credenciales para Drive
    drive_creds = auth_service.get_credentials(for_service="drive")
    drive_service = GoogleDriveService(drive_creds, config)
    print("✅ Google Drive Service inicializado")

    # Credenciales para Gmail
    try:
        gmail_creds = auth_service.get_credentials(for_service="gmail")
        gmail_service = GmailService(gmail_creds, config)
        print("✅ Gmail Service inicializado")
    except Exception as e:
        print(f"⚠️  Gmail Service no disponible: {e}")
        gmail_service = None

    return drive_service, gmail_service


def ejemplo_1_crear_carpeta_con_placeholders():
    """
    EJEMPLO 1: Crear carpeta de cliente con archivos placeholder
    """
    print("\n" + "=" * 60)
    print("EJEMPLO 1: Crear carpeta con 4 archivos placeholder")
    print("=" * 60)

    config = AppConfig()
    drive_service, _ = setup_servicios(config)

    # Datos del cliente (pueden venir de cualquier fuente)
    cliente = Cliente(nombre="DUANY", apellido1="BARO", apellido2="MENÉNDEZ")

    # Crear 4 archivos placeholder (2 imágenes + 2 fotos)
    print("\n📦 Generando archivos placeholder...")
    archivos = [
        PlaceholderGenerator.crear_imagen_placeholder("imagen1.png", color="blue"),
        PlaceholderGenerator.crear_imagen_placeholder("imagen2.png", color="green"),
        PlaceholderGenerator.crear_imagen_placeholder("foto1.png", color="red"),
        PlaceholderGenerator.crear_imagen_placeholder("foto2.png", color="purple"),
    ]

    # Crear carpeta y subir archivos
    folder_id, archivos_subidos = drive_service.crear_carpeta_con_archivos(
        nombre_carpeta=cliente.nombre_carpeta,
        archivos=archivos,
        parent_id=config.drive_root_folder_id,
    )

    print(f"\n✅ COMPLETADO")
    print(f"   Carpeta ID: {folder_id}")
    print(f"   Archivos: {list(archivos_subidos.keys())}")


def ejemplo_2_job_email_a_drive():
    """
    EJEMPLO 2: Job que procesa correos y sube Excel a Drive
    """
    print("\n" + "=" * 60)
    print("EJEMPLO 2: Procesar correos y subir a Drive")
    print("=" * 60)

    config = AppConfig()
    drive_service, gmail_service = setup_servicios(config)

    if not gmail_service:
        print("❌ Gmail service no disponible")
        return

    # Crear job
    job = EmailProcessorJob(gmail_service, drive_service, config)

    # Opción 1: Ejecutar una vez
    resultados = job.procesar_correos_nuevos(
        guardar_local=True,
        subir_a_drive=True,
        folder_id_drive=config.drive_root_folder_id,
    )

    print(f"\n✅ Procesados {len(resultados)} correos")

    for correo, file_ids in resultados:
        print(f"\n  📧 {correo.subject}")
        print(f"     Archivos en Drive: {len(file_ids)}")
        for file_id in file_ids:
            print(f"       - {file_id}")

    # Opción 2: Loop continuo (descomentar para usar)
    # job.ejecutar_loop(
    #     guardar_local=True,
    #     subir_a_drive=True,
    #     folder_id_drive=config.drive_root_folder_id
    # )


def ejemplo_3_monitorear_y_validar():
    """
    EJEMPLO 3: Monitorear cambios en Excel y ejecutar validación
    """
    print("\n" + "=" * 60)
    print("EJEMPLO 3: Monitorear cambios con validación automática")
    print("=" * 60)

    config = AppConfig()
    drive_service, _ = setup_servicios(config)

    # Crear job de monitoreo
    job = DriveMonitorJob(drive_service, config)

    # ID de carpeta a monitorear
    folder_id = config.drive_root_folder_id or "ID_DE_TU_CARPETA"

    print(f"\n📂 Monitoreando carpeta: {folder_id}")

    # Opción 1: Chequear una vez con callback de validación
    cambios = job.procesar_cambios(
        folder_id=folder_id, callback_on_change=ejemplo_callback_validacion
    )

    print(f"\n✅ Detectados {len(cambios)} cambios")

    # Opción 2: Loop continuo con callback (descomentar para usar)
    # job.ejecutar_loop(
    #     folder_id=folder_id,
    #     callback_on_change=ejemplo_callback_validacion
    # )


def ejemplo_4_leer_y_actualizar_excel():
    """
    EJEMPLO 4: Leer Excel desde Drive, modificar y actualizar
    """
    print("\n" + "=" * 60)
    print("EJEMPLO 4: Leer y actualizar Excel en Drive")
    print("=" * 60)

    config = AppConfig()
    drive_service, _ = setup_servicios(config)

    # Primero, crear un Excel de prueba
    print("\n📝 Creando Excel de prueba...")
    excel_placeholder = PlaceholderGenerator.crear_excel_placeholder("test_excel.xlsx")

    folder_id = config.drive_root_folder_id or "ID_DE_TU_CARPETA"
    file_id = drive_service.subir_archivo(excel_placeholder, folder_id)

    print(f"   Excel creado con ID: {file_id}")

    # Leer el Excel
    print("\n📖 Leyendo Excel desde Drive...")
    excel_data = drive_service.leer_excel_desde_drive(file_id)

    print(f"   Hojas encontradas: {excel_data.sheet_names}")
    print(f"   Total filas: {sum(len(rows) for rows in excel_data.data.values())}")

    # Modificar datos
    print("\n✏️  Modificando datos...")
    for sheet_name, rows in excel_data.data.items():
        df = pd.DataFrame(rows)

        # Ejemplo: Agregar una nueva columna
        df["Total"] = df["Cantidad"] * df["Precio"]

        # Ejemplo: Agregar una fila resumen
        resumen = pd.DataFrame(
            [
                {
                    "ID": "TOTAL",
                    "Nombre": "",
                    "Cantidad": df["Cantidad"].sum(),
                    "Precio": df["Precio"].mean(),
                    "Total": df["Total"].sum(),
                }
            ]
        )

        df = pd.concat([df, resumen], ignore_index=True)

        # Actualizar en Drive
        print(f"\n💾 Actualizando hoja '{sheet_name}' en Drive...")
        drive_service.actualizar_excel_en_drive(
            file_id=file_id, dataframes={sheet_name: df}
        )

    print("\n✅ Excel actualizado correctamente")

    # Verificar cambios
    print("\n🔍 Verificando cambios...")
    excel_data_updated = drive_service.leer_excel_desde_drive(file_id)

    for sheet_name, rows in excel_data_updated.data.items():
        df = pd.DataFrame(rows)
        print(f"\n  📋 Hoja '{sheet_name}':")
        print(f"     Columnas: {list(df.columns)}")
        print(f"     Filas: {len(df)}")
        if "Total" in df.columns:
            print(f"     ✅ Nueva columna 'Total' agregada")


def ejemplo_5_flujo_completo():
    """
    EJEMPLO 5: Flujo completo - Email -> Drive -> Monitor -> Validación
    """
    print("\n" + "=" * 60)
    print("EJEMPLO 5: Flujo completo integrado")
    print("=" * 60)

    config = AppConfig()
    drive_service, gmail_service = setup_servicios(config)

    if not gmail_service:
        print("⚠️  Modo sin Gmail - Usando Excel de prueba")

        # Crear Excel de prueba directamente en Drive
        excel_placeholder = PlaceholderGenerator.crear_excel_placeholder(
            "flujo_test.xlsx"
        )
        folder_id = config.drive_root_folder_id or "ID_DE_TU_CARPETA"
        file_id = drive_service.subir_archivo(excel_placeholder, folder_id)
        print(f"✅ Excel de prueba creado: {file_id}")

    else:
        # Paso 1: Procesar correos y subir a Drive
        print("\n📧 PASO 1: Procesar correos...")
        email_job = EmailProcessorJob(gmail_service, drive_service, config)
        resultados = email_job.procesar_correos_nuevos(
            guardar_local=True,
            subir_a_drive=True,
            folder_id_drive=config.drive_root_folder_id,
        )

        if resultados:
            print(f"✅ {len(resultados)} correos procesados")
        else:
            print("ℹ️  No hay correos nuevos")

    # Paso 2: Monitorear cambios
    print("\n📊 PASO 2: Monitorear cambios en Drive...")
    monitor_job = DriveMonitorJob(drive_service, config)

    folder_id = config.drive_root_folder_id or "ID_DE_TU_CARPETA"

    cambios = monitor_job.procesar_cambios(
        folder_id=folder_id, callback_on_change=ejemplo_callback_validacion
    )

    if cambios:
        print(f"✅ {len(cambios)} archivos modificados procesados")
    else:
        print("ℹ️  No hay cambios detectados")

    print("\n✅ FLUJO COMPLETO FINALIZADO")


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    print(
        """
╔════════════════════════════════════════════════════════════════╗
║   Sistema de Gestión Google Drive y Gmail                     ║
║   Con Service Accounts y Pydantic                             ║
╚════════════════════════════════════════════════════════════════╝

📚 EJEMPLOS DISPONIBLES:

1. ejemplo_1_crear_carpeta_con_placeholders()
   - Crea carpeta de cliente con 4 archivos placeholder

2. ejemplo_2_job_email_a_drive()
   - Procesa correos y sube Excel adjuntos a Drive

3. ejemplo_3_monitorear_y_validar()
   - Monitorea cambios en Excel y ejecuta validaciones

4. ejemplo_4_leer_y_actualizar_excel()
   - Lee Excel desde Drive, modifica y actualiza

5. ejemplo_5_flujo_completo()
   - Flujo completo integrado: Email -> Drive -> Monitor

⚙️  CONFIGURACIÓN:
   - Edita .env con tus credenciales
   - Coloca service-account.json en la raíz
   - Configura DRIVE_ROOT_FOLDER_ID y filtros Gmail

🚀 Ejecuta un ejemplo descomentando la línea correspondiente:
    """
    )

    # Ejecutar ejemplo (descomentar el que necesites)
    # ejemplo_1_crear_carpeta_con_placeholders()
    # ejemplo_2_job_email_a_drive()
    # ejemplo_3_monitorear_y_validar()
    # ejemplo_4_leer_y_actualizar_excel()
    # ejemplo_5_flujo_completo()
